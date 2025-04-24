import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font

def n11_komisyon_orani_duzenle():
    """
    N11 Pazar Yerleri Kategori Komisyon Oran Şablonu.xlsx dosyasındaki
    Komisyon Oranı sütunundaki değerlerin başına % işareti ekler.
    
    Returns:
        tuple: (bool, str) - İşlemin başarılı olup olmadığı ve sonuç mesajı
    """
    try:
        # Excel dosyasının yolu
        input_file = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))), 
            "documents", "Pazar Yerleri Komisyon Oranları", 
            "(N11) Pazar Yerleri Kategori Komisyon Oran Şablonu.xlsx"
        )
        
        # Çıktı dosyasının yolu
        output_file = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))), 
            "documents", "Pazar Yerleri Komisyon Oranları", 
            "(N11) Pazar Yerleri Kategori Komisyon Oran Şablonu (Düzenlenmiş).xlsx"
        )
        
        print(f"Excel dosyası yolu: {input_file}")
        
        # Excel dosyasını pandas ile oku
        df = pd.read_excel(input_file)
        
        # Excel verilerindeki sütun adlarını kontrol et
        print(f"Excel sütun adları: {df.columns.tolist()}")
        
        # Komisyon Oranı sütununu bul
        komisyon_sutunu = None
        for col in df.columns:
            if "Komisyon" in col and "Oran" in col:
                komisyon_sutunu = col
                break
        
        if not komisyon_sutunu:
            return False, "Komisyon Oranı sütunu bulunamadı."
        
        print(f"Komisyon Oranı sütunu: {komisyon_sutunu}")
        
        # Veri yapısını kontrol etmek için birkaç satır göster
        print("İlk 5 satır:")
        print(df.head())
        
        # Komisyon oranlarını düzenle
        for idx, row in df.iterrows():
            komisyon_degeri = row[komisyon_sutunu]
            
            # Boş değerleri atla
            if pd.isna(komisyon_degeri):
                continue
            
            # Eğer değer zaten % ile başlıyorsa, değiştirme
            if isinstance(komisyon_degeri, str) and komisyon_degeri.startswith('%'):
                continue
            
            # Değeri string'e çevir ve % işareti ekle
            if isinstance(komisyon_degeri, (int, float)):
                df.at[idx, komisyon_sutunu] = f"%{komisyon_degeri}"
            elif isinstance(komisyon_degeri, str):
                # Eğer zaten % işareti yoksa ekle
                if not komisyon_degeri.startswith('%'):
                    df.at[idx, komisyon_sutunu] = f"%{komisyon_degeri}"
        
        # Düzenlenmiş verileri yeni Excel dosyasına kaydet
        df.to_excel(output_file, index=False)
        
        print(f"Düzenlenmiş veriler şu dosyaya kaydedildi: {output_file}")
        
        # Excel dosyasını openpyxl ile aç ve formatlamayı düzenle
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # Komisyon Oranı sütununu bul
        komisyon_sutun_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if "Komisyon" in str(cell.value) and "Oran" in str(cell.value):
                komisyon_sutun_idx = idx
                break
        
        if komisyon_sutun_idx:
            # Komisyon Oranı sütunundaki tüm hücreleri % formatına çevir
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=komisyon_sutun_idx)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('%'):
                    # Hücreyi % formatına çevir
                    cell.number_format = '0.00%'
                    # % işaretini kaldır ve sayıya çevir
                    try:
                        value = float(cell.value.replace('%', '')) / 100
                        cell.value = value
                    except ValueError:
                        pass
        
        # Düzenlenmiş Excel dosyasını kaydet
        wb.save(output_file)
        
        return True, f"Komisyon oranları başarıyla düzenlendi ve {output_file} dosyasına kaydedildi."
        
    except Exception as e:
        import traceback
        print(f"Hata oluştu: {str(e)}")
        print(traceback.format_exc())
        return False, f"Hata: {str(e)}"

if __name__ == "__main__":
    basarili, mesaj = n11_komisyon_orani_duzenle()
    print(mesaj)
    if not basarili:
        sys.exit(1)  # Hata durumunda 1 döndür
    sys.exit(0)  # Başarılı durumda 0 döndür 